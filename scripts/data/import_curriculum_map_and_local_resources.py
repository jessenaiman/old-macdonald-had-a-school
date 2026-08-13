#!/usr/bin/env python3
"""Import the curriculum map and the user-supplied local resource workbook.

This is intentionally a small, idempotent batch. Exact duplicate resources are
reused; they do not prevent new map rows or new resource links from being
processed.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


BATCH_ID = "data-20260813-curriculum-map-local-resources-v3"
GRADE_KEYS = {"Grade 1": "grade-1", "Grade 2": "grade-2", "Grade 3": "grade-3"}
MAP_HEADERS = ("Week", "Month", "Strand", "Lesson Name", "Ontario Code")
RESOURCE_HEADERS = (
    "Resource Name", "Type", "Creator/Artist", "URL Link", "Age Band",
    "Curriculum Categories", "Tags", "Verified Y/N", "Verified By",
    "Verified Date", "Notes",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


NEW_TOPIC_METADATA = {
    normalize("Ask and answer questions about key details in a text read aloud"): {
        "category": "Speaking & Listening",
        "teacher_title": "Read-Aloud Detectives: Find the Key Details",
        "teacher_summary": "Students listen to a short read-aloud, identify the people, actions, or facts that matter most, and support an answer by pointing back to what they heard. The teacher can use the responses to decide whether the next lesson should revisit listening for detail or move toward retelling.",
    }
}


def project_path(path: Path, root: Path) -> str:
    return path.resolve().relative_to(root.resolve()).as_posix()


def normalized_headers(values: tuple[object, ...]) -> tuple[str, ...]:
    return tuple(normalize(value) for value in values)


def validate_workbooks(map_path: Path, resources_path: Path) -> dict[str, object]:
    """Fail before opening SQLite if an earlier workbook shape is supplied."""
    map_workbook = load_workbook(map_path, read_only=True, data_only=True)
    resource_workbook = load_workbook(resources_path, read_only=True, data_only=True)
    try:
        missing_map_sheets = [name for name in GRADE_KEYS if name not in map_workbook.sheetnames]
        if missing_map_sheets:
            raise ValueError(f"Curriculum map is missing sheets: {', '.join(missing_map_sheets)}")
        for name in GRADE_KEYS:
            headers = normalized_headers(tuple(next(map_workbook[name].iter_rows(max_row=1, values_only=True))))
            if headers != normalized_headers(MAP_HEADERS):
                raise ValueError(f"Unexpected headers in {name}: {headers!r}")
        if "Resources" not in resource_workbook.sheetnames:
            raise ValueError("Resource workbook is missing the Resources sheet")
        headers = normalized_headers(tuple(next(resource_workbook["Resources"].iter_rows(max_row=1, values_only=True))))
        if headers != normalized_headers(RESOURCE_HEADERS):
            raise ValueError(f"Unexpected headers in Resources: {headers!r}")
        return {
            "map_sheets": {name: map_workbook[name].max_row - 1 for name in GRADE_KEYS},
            "resource_rows": resource_workbook["Resources"].max_row - 1,
        }
    finally:
        map_workbook.close()
        resource_workbook.close()


def source_document_id(db: sqlite3.Connection, path: str, checksum: str) -> int:
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        """
        INSERT INTO source_documents (source_path, source_kind, review_state, checksum, imported_at)
        VALUES (?, 'xlsx', 'reviewed', ?, ?)
        ON CONFLICT(source_path) DO UPDATE SET
          review_state = 'reviewed', checksum = excluded.checksum, imported_at = excluded.imported_at
        """,
        (path, checksum, now),
    )
    return db.execute("SELECT id FROM source_documents WHERE source_path = ?", (path,)).fetchone()[0]


def source_id(db: sqlite3.Connection, path: str, checksum: str) -> int:
    row = db.execute("SELECT id FROM sources WHERE path_or_url = ?", (path,)).fetchone()
    if row:
        db.execute("UPDATE sources SET checksum = ?, type = 'xlsx' WHERE id = ?", (checksum, row[0]))
        return row[0]
    cursor = db.execute(
        "INSERT INTO sources (path_or_url, type, checksum) VALUES (?, 'xlsx', ?)",
        (path, checksum),
    )
    return cursor.lastrowid


def import_map(db: sqlite3.Connection, workbook_path: Path, root: Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    grade_ids = {
        row[0]: row[1]
        for row in db.execute("SELECT key, id FROM grades WHERE key IN ('grade-1', 'grade-2', 'grade-3')")
    }
    topic_grade_rows = db.execute(
        """
        SELECT tg.id, tg.topic_id, tg.grade_id, t.topic
        FROM topic_grades tg
        JOIN topics t ON t.id = tg.topic_id
        WHERE tg.grade_id IN (?, ?, ?)
        """,
        (grade_ids["grade-1"], grade_ids["grade-2"], grade_ids["grade-3"]),
    ).fetchall()
    topic_grades = {(grade_id, normalize(topic)): (tg_id, topic_id) for tg_id, topic_id, grade_id, topic in topic_grade_rows}

    stats = {"map_rows": 0, "topics_inserted": 0, "pacing_inserted": 0, "pacing_updated": 0, "stale_earlier_placements_removed": 0, "standards_linked": 0, "unresolved": 0}
    source_path = project_path(workbook_path, root)
    for sheet_name, grade_key in GRADE_KEYS.items():
        worksheet = workbook[sheet_name]
        for week, month, strand, lesson_name, ontario_code in worksheet.iter_rows(min_row=2, values_only=True):
            if not lesson_name or not month or not week:
                continue
            stats["map_rows"] += 1
            grade_id = grade_ids[grade_key]
            match = topic_grades.get((grade_id, normalize(lesson_name)))
            if not match:
                metadata = NEW_TOPIC_METADATA.get(normalize(lesson_name))
                subject = db.execute("SELECT id FROM subjects WHERE label = ?", (strand,)).fetchone()
                if not metadata or not subject:
                    stats["unresolved"] += 1
                    continue
                cursor = db.execute(
                    """
                    INSERT INTO topics
                      (subject_id, category, topic, skill, sequence, taught_status,
                       teacher_title, teacher_summary, teacher_title_state)
                    VALUES (?, ?, ?, ?, ?, '', ?, ?, 'editorial')
                    """,
                    (subject[0], metadata["category"], lesson_name, lesson_name, float(week), metadata["teacher_title"], metadata["teacher_summary"]),
                )
                topic_id = cursor.lastrowid
                cursor = db.execute(
                    "INSERT INTO topic_grades (topic_id, grade_id) VALUES (?, ?)",
                    (topic_id, grade_id),
                )
                topic_grade_id = cursor.lastrowid
                topic_grades[(grade_id, normalize(lesson_name))] = (topic_grade_id, topic_id)
                match = (topic_grade_id, topic_id)
                stats["topics_inserted"] += 1
            topic_grade_id, topic_id = match
            map_note = f"Source: {source_path}; strand={strand or 'not recorded'}; Ontario code={ontario_code or 'not recorded'}"
            existing = db.execute(
                "SELECT id, week_number, month FROM weekly_pacing WHERE topic_grade_id = ? ORDER BY week_number",
                (topic_grade_id,),
            ).fetchall()
            same_slot = next((row for row in existing if row[1] == week and row[2] == month), None)
            if same_slot:
                db.execute("UPDATE weekly_pacing SET notes = ? WHERE id = ?", (map_note, same_slot[0]))
                pacing_id = same_slot[0]
            elif len(existing) == 1:
                pacing_id = existing[0][0]
                db.execute(
                    "UPDATE weekly_pacing SET week_number = ?, month = ?, notes = ? WHERE id = ?",
                    (week, month, map_note, pacing_id),
                )
                stats["pacing_updated"] += 1
            else:
                cursor = db.execute(
                    "INSERT INTO weekly_pacing (topic_grade_id, week_number, month, notes) VALUES (?, ?, ?, ?)",
                    (topic_grade_id, week, month, map_note),
                )
                pacing_id = cursor.lastrowid
                stats["pacing_inserted"] += 1

            removed = db.execute(
                """
                DELETE FROM weekly_pacing
                WHERE topic_grade_id = ?
                  AND week_number < ?
                  AND COALESCE(notes, '') NOT LIKE ?
                """,
                (topic_grade_id, week, f"Source: {source_path}%"),
            ).rowcount
            stats["stale_earlier_placements_removed"] += removed

            standard = db.execute(
                "SELECT id, code FROM standards WHERE framework = 'Ontario' AND full_text = ? ORDER BY id LIMIT 1",
                (lesson_name,),
            ).fetchone()
            if standard:
                standard_id = standard[0]
                if not standard[1] and ontario_code:
                    db.execute("UPDATE standards SET code = ? WHERE id = ?", (ontario_code, standard_id))
            else:
                cursor = db.execute(
                    """
                    INSERT INTO standards (framework, code, full_text, source)
                    VALUES ('Ontario', ?, ?, ?)
                    """,
                    (ontario_code, lesson_name, "Curriculum_Map.xlsx"),
                )
                standard_id = cursor.lastrowid
            linked = db.execute(
                "SELECT 1 FROM topic_standards WHERE topic_id = ? AND standard_id = ?",
                (topic_id, standard_id),
            ).fetchone()
            if not linked:
                db.execute(
                    "INSERT INTO topic_standards (topic_id, standard_id, alignment_notes) VALUES (?, ?, ?)",
                    (topic_id, standard_id, map_note),
                )
                stats["standards_linked"] += 1
    return stats


def import_resources(db: sqlite3.Connection, workbook_path: Path, root: Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["Resources"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    source_path = project_path(workbook_path, root)
    workbook_source_id = source_id(db, source_path, sha256(workbook_path))
    stats = {"resource_rows": 0, "resources_inserted": 0, "resources_reused": 0, "materials_linked": 0}
    resources: dict[str, int] = {}
    # These are explicit reviewed URL overrides for the three local workbook
    # rows; they are not a general web lookup or a duplicate-import rule.
    verified_urls = {
        normalize("Old MacDonald Had a Farm"): "https://www.teachingenglish.org.uk/sites/teacheng/files/D150%20Teacher%20Notes_Old%20MacDonald%20v4.pdf",
        normalize("Open Shut Them"): "https://supersimple.com/song/open-shut-them/",
        normalize("The Very Hungry Caterpillar"): "https://eric-carle.com/eric-carle-book-gallery/the-very-hungry-caterpillar-1969/",
    }
    for name, resource_type, creator, url, age_band, categories, tags, verified, verified_by, verified_date, notes in rows:
        if not name:
            continue
        stats["resource_rows"] += 1
        existing = None
        if url:
            existing = db.execute("SELECT id FROM resources WHERE url = ? ORDER BY id LIMIT 1", (url,)).fetchone()
        if not existing:
            existing = db.execute(
                "SELECT id FROM resources WHERE lower(name) = lower(?) ORDER BY id LIMIT 1", (name,)
            ).fetchone()
        resource_kind = "web"
        source_tags = str(tags or "")
        if normalize(name) == normalize("The Very Hungry Caterpillar"):
            source_tags = ", ".join(
                tag.strip()
                for tag in source_tags.split(",")
                if tag.strip() and normalize(tag) != normalize("Old MacDonald")
            )
        if normalize(name) == normalize("Old MacDonald Had a Farm"):
            source_notes = "Online verification identifies this as a traditional song; do not infer public-domain status for a particular recording or arrangement."
        elif normalize(name) == normalize("Open Shut Them"):
            source_notes = "Online verification confirms an official Super Simple teaching page with gestures, lyrics, flashcards, and a worksheet."
        elif normalize(name) == normalize("The Very Hungry Caterpillar"):
            source_notes = "Online verification confirms the official Eric Carle page describes days of the week, counting, nutrition, and the caterpillar life cycle."
        else:
            source_notes = None
        description = "; ".join(
            value for value in [
                f"Creator/artist: {creator}" if creator else None,
                f"Workbook resource type: {resource_type}" if resource_type else None,
                f"Age band: {age_band}" if age_band else None,
                f"Categories: {categories}" if categories else None,
                f"Tags: {source_tags}" if source_tags else None,
                f"Verified by: {verified_by} ({verified_date})" if verified_by or verified_date else None,
                str(notes) if notes else None,
                source_notes,
            ]
            if value
        )
        verified_value = 1 if str(verified or "").casefold() in {"y", "yes", "true", "1"} else 0
        if existing:
            resource_id = existing[0]
            db.execute(
                "UPDATE resources SET description = ?, url = ?, verified = ?, source_id = ? WHERE id = ?",
                (description, verified_urls.get(normalize(name), url), verified_value, workbook_source_id, resource_id),
            )
            stats["resources_reused"] += 1
        else:
            cursor = db.execute(
                """
                INSERT INTO resources (name, type, description, url, verified, source_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (name, resource_kind, description, verified_urls.get(normalize(name), url), verified_value, workbook_source_id),
            )
            resource_id = cursor.lastrowid
            stats["resources_inserted"] += 1
        resources[normalize(name)] = resource_id

    links = [
        (
            "Old MacDonald Had a Farm",
            427,
            "focus",
            "opening",
            "circle-time-core",
            "The workbook identifies this traditional song for infant-to-preschool use and links it to language/vocabulary and science/nature; use the existing Old MacDonald topic as the direct lesson home without creating another song record.",
        ),
        (
            "Open Shut Them",
            404,
            "supporting",
            "opening",
            "circle-time-core",
            "The workbook identifies this fingerplay for infant-to-preschool use and categorizes it for fine-motor and classroom-routine work; offer it as a brief hand warm-up before palmar-grasp exploration.",
        ),
        (
            "The Very Hungry Caterpillar",
            411,
            "focus",
            "guided-practice",
            "circle-time-core",
            "The workbook identifies this book for toddler-to-preschool use across literacy, science, and mathematics; use it as an optional picture-book choice for pointing, shared attention, and oral retell.",
        ),
    ]
    for name, topic_id, role, phase, routine, rationale in links:
        resource_id = resources.get(normalize(name))
        if resource_id is None:
            continue
        before = db.execute(
            "SELECT 1 FROM topic_materials WHERE topic_id = ? AND material_kind = 'resource' AND material_id = ?",
            (topic_id, resource_id),
        ).fetchone()
        db.execute(
            """
            INSERT INTO topic_materials
              (topic_id, material_kind, material_id, role, use_in_phase, routine_slot, teacher_rationale)
            VALUES (?, 'resource', ?, ?, ?, ?, ?)
            ON CONFLICT(topic_id, material_kind, material_id) DO UPDATE SET
              role = excluded.role, use_in_phase = excluded.use_in_phase,
              routine_slot = excluded.routine_slot, teacher_rationale = excluded.teacher_rationale
            """,
            (topic_id, resource_id, role, phase, routine, rationale),
        )
        if not before:
            stats["materials_linked"] += 1
    return stats


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--map", dest="map_path", type=Path, default=Path("data/Curriculum_Map.xlsx"))
    parser.add_argument("--resources", dest="resources_path", type=Path, default=Path("data/Curriculum_Resource_Database.xlsx"))
    parser.add_argument("--db", type=Path, default=Path("data/omhas.db"))
    parser.add_argument("--apply", action="store_true")
    parser.add_argument(
        "--batch-id",
        default=BATCH_ID,
        help="Explicit revision identifier; use a new value only after reviewing source changes.",
    )
    args = parser.parse_args()
    root = Path.cwd().resolve()
    map_path = args.map_path.resolve()
    resources_path = args.resources_path.resolve()
    if not map_path.is_file() or not resources_path.is_file():
        raise SystemExit("Both workbook paths must exist before importing.")
    workbook_preview = validate_workbooks(map_path, resources_path)
    db = sqlite3.connect(args.db)
    db.execute("PRAGMA foreign_keys = ON")
    try:
        if db.execute("SELECT 1 FROM import_batches WHERE migration_id = ?", (args.batch_id,)).fetchone():
            print(json.dumps({"status": "already applied", "batch_id": args.batch_id}))
            return 0
        result = {"batch_id": args.batch_id, "status": "dry-run", "workbook_preview": workbook_preview}
        if args.apply:
            db.execute("BEGIN")
            map_path_rel = project_path(map_path, root)
            resources_path_rel = project_path(resources_path, root)
            source_document_id(db, map_path_rel, sha256(map_path))
            source_document_id(db, resources_path_rel, sha256(resources_path))
            result["map"] = import_map(db, map_path, root)
            result["resources"] = import_resources(db, resources_path, root)
            db.execute(
                "INSERT INTO schema_migrations (migration_id, applied_at, omhas_sha256, curriculum_sha256, generated_sha256) VALUES (?, CURRENT_TIMESTAMP, 'DATA_IMPORT', 'NOT_APPLICABLE', 'NOT_APPLICABLE')",
                (args.batch_id,),
            )
            db.execute(
                "INSERT INTO import_batches (migration_id, source_name, source_path, source_sha256, imported_at) VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)",
                (args.batch_id, "curriculum_map_and_local_resources", f"{map_path_rel}; {resources_path_rel}", sha256(map_path) + ";" + sha256(resources_path)),
            )
            db.execute("COMMIT")
            result["status"] = "applied"
        else:
            result["map_source"] = project_path(map_path, root)
            result["resource_source"] = project_path(resources_path, root)
        print(json.dumps(result, ensure_ascii=False))
        return 0
    except Exception:
        if db.in_transaction:
            db.execute("ROLLBACK")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
