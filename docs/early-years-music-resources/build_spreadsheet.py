"""Build Song_Resources_Review.xlsx from extracted records.
Applies: uark duplicate removal, Oxford preview noise filter,
rough-quality filter, then writes the review spreadsheet."""
import json, os, re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = r"C:\Users\jesse\OneDrive\Documents\New project\data\sources\early-years-music-resources"
SRC = os.path.join(ROOT, "metadata", "_extracted_songs.json")
OUT = os.path.join(ROOT, "Song_Resources_Review.xlsx")

with open(SRC, encoding="utf-8") as fh:
    data = json.load(fh)

# ---- dedupe: uark-ecep (01) and uark-ecec (02) are byte-identical ----
data["02-educators-publishers"] = [r for r in data["02-educators-publishers"]
                                    if r["source_file"] != "uark-ecec-favorite-songs-and-fingerplays.pdf"]

# ---- Oxford Kodaly preview: keep only lesson/song content, drop TOC/author/front-matter ----
ox_noise = {"micheál houlahan and philip tacka", "philip tacka", "oxford university press",
            "houlahan, micheál.", "houlahan, philip tacka.", "contents", "purpose of book",
            "how is this book different from other related texts?", "what are the connections between music literacy"}
kept_ox = []
for r in data["02-educators-publishers"]:
    if r["source_file"] != "oxford-kodaly-in-kindergarten-classroom-preview.pdf":
        kept_ox.append(r)
        continue
    t = r["song_title"].strip().lower().rstrip(".")
    if t in ox_noise or re.match(r"^3\.\s*kindergarten", t) or t.startswith("south korea"):
        continue
    if r["extraction_quality"] == "rough":
        continue
    kept_ox.append(r)
data["02-educators-publishers"] = kept_ox

# ---- drop remaining rough blocks with thin lyrics ----
for cat in data:
    data[cat] = [r for r in data[cat]
                 if not (r["extraction_quality"] == "rough" and len(re.findall(r'[A-Za-z]{3,}', r["lyrics"])) < 20)]

# ---- flatten ----
CAT_LABEL = {
    "01-libraries-agencies": "Libraries & Agencies",
    "02-educators-publishers": "Educators & Publishers",
    "03-performers-programs": "Performers & Programs",
}
all_recs = []
for cat, recs in data.items():
    for r in recs:
        all_recs.append({
            "song_title": r.get("song_title", ""),
            "actions": r.get("actions", ""),
            "lyrics": r.get("lyrics", ""),
            "source_file": r.get("source_file", ""),
            "source_title": r.get("source_title", ""),
            "creator": r.get("creator", ""),
            "age_range": r.get("age_range", ""),
            "url": r.get("url", ""),
            "category": CAT_LABEL.get(cat, cat),
            "quality": r.get("extraction_quality", ""),
            "notes": r.get("notes", ""),
        })
print(f"Final record count: {len(all_recs)}")
print("by category:", dict(Counter(r["category"] for r in all_recs)))
print("by quality:", dict(Counter(r["quality"] for r in all_recs)))

# ---- build workbook ----
wb = Workbook()
HDR_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
QUALITY_FILL = {
    "clean": PatternFill("solid", fgColor="E2EFDA"),
    "ok": PatternFill("solid", fgColor="FFF2CC"),
    "numbered": PatternFill("solid", fgColor="DDEBF7"),
    "rough": PatternFill("solid", fgColor="FCE4EC"),
}

ws = wb.active
ws.title = "Songs"
headers = ["#", "Song / Fingerplay / Rhyme", "Actions / Lessons", "Lyrics",
           "Source File", "Source Title", "Creator / Organization", "Age Range",
           "URL", "Category", "Extraction Quality", "Notes"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = BORDER

for i, r in enumerate(all_recs, 1):
    ws.append([i, r["song_title"], r["actions"], r["lyrics"],
               r["source_file"], r["source_title"], r["creator"], r["age_range"],
               r["url"], r["category"], r["quality"], r["notes"]])
    row = i + 1
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT; cell.border = BORDER
        cell.alignment = WRAP if c in (2, 3, 4, 6, 7, 9) else Alignment(vertical="top")
    qf = QUALITY_FILL.get(r["quality"])
    if qf:
        ws.cell(row=row, column=11).fill = qf

for c, w in enumerate([6, 30, 40, 60, 34, 34, 30, 20, 45, 22, 12, 16], 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{len(all_recs)+1}"

# ---- Sources sheet ----
ws2 = wb.create_sheet("Sources")
src_rows = {}
for r in all_recs:
    key = r["source_file"]
    d = src_rows.setdefault(key, {"title": r["source_title"], "creator": r["creator"],
                                   "age": r["age_range"], "url": r["url"], "category": r["category"],
                                   "count": 0, "clean": 0, "ok": 0, "numbered": 0, "rough": 0})
    d["count"] += 1
    d[r["quality"]] = d.get(r["quality"], 0) + 1

ws2.append(["Source File", "Source Title", "Creator / Organization", "Age Range",
            "Category", "Songs Extracted", "Clean", "OK", "Numbered", "Rough", "URL"])
for c in range(1, 12):
    cell = ws2.cell(row=1, column=c)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = BORDER

for fname, d in sorted(src_rows.items()):
    ws2.append([fname, d["title"], d["creator"], d["age"], d["category"], d["count"],
                d["clean"], d["ok"], d["numbered"], d["rough"], d["url"]])
    row = ws2.max_row
    for c in range(1, 12):
        cell = ws2.cell(row=row, column=c)
        cell.font = BODY_FONT; cell.border = BORDER
        if c in (2, 3, 11):
            cell.alignment = WRAP

for c, w in enumerate([38, 40, 32, 22, 24, 12, 8, 8, 10, 8, 50], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:K{ws2.max_row}"

# ---- Image-based sheet ----
ws3 = wb.create_sheet("Image-Based (No Text)")
image_files = [
    ("swanton-public-library-bounce-rhymes.pdf", "Swanton Public Library, VT"),
    ("sc-school-kindergarten-music-movement-activities.pdf", "SC school (unidentified)"),
    ("jack-hartmann-a-to-z-flash-cards.pdf", "Jack Hartmann / Hop 2 It Music"),
    ("jack-hartmann-stretchy-word-snake-coloring-activity.pdf", "Jack Hartmann / Hop 2 It Music"),
    ("mother-goose-club-five-little-monkeys-activity-book.pdf", "Mother Goose Club"),
    ("mother-goose-club-five-little-monkeys-lyric-book.pdf", "Mother Goose Club"),
    ("mother-goose-club-itsy-bitsy-spider-lyric-book.pdf", "Mother Goose Club"),
    ("raffi-baby-beluga-lyrics-arrangement.pdf", "Raffi"),
    ("raffi-down-by-the-bay-lyrics-arrangement.pdf", "Raffi"),
    ("raffi-wheels-on-the-bus-lyrics-arrangement.pdf", "Raffi"),
    ("the-learning-station-tony-chestnut-activity-handout.pdf", "The Learning Station"),
    ("the-wiggles-dice-roll-activity.pdf", "The Wiggles"),
    ("the-wiggles-wiggle-and-learn-circus-colouring.pdf", "The Wiggles"),
]
ws3.append(["Source File", "Creator / Organization", "Status"])
for c in range(1, 4):
    cell = ws3.cell(row=1, column=c)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
for fname, creator in image_files:
    ws3.append([fname, creator, "Image-based PDF — no text layer; lyrics/actions not machine-extractable. Needs OCR or manual entry."])
    for c in range(1, 4):
        cell = ws3.cell(row=ws3.max_row, column=c)
        cell.font = BODY_FONT; cell.border = BORDER; cell.alignment = WRAP
ws3.column_dimensions["A"].width = 55
ws3.column_dimensions["B"].width = 32
ws3.column_dimensions["C"].width = 80
ws3.freeze_panes = "A2"

wb.save(OUT)
print("saved:", OUT)
