from pathlib import Path
from openpyxl import load_workbook

root = Path(r"C:\Users\jesse\OneDrive\Documents\New project\resources")
for path in sorted(root.rglob("*.xlsx")):
    print(f"\nWORKBOOK {path.relative_to(root)}")
    workbook = load_workbook(path, read_only=True, data_only=False)
    for sheet in workbook.worksheets:
        print(f"  SHEET {sheet.title!r} rows={sheet.max_row} cols={sheet.max_column}")
        shown = 0
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [value for value in row]
            if any(value not in (None, "") for value in values):
                print(f"    row {row_number}: {values[:18]}")
                shown += 1
                if shown >= 3:
                    break
    workbook.close()
